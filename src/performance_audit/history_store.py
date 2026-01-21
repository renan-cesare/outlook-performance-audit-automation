
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_HEADERS = [
    "datetime_sent",
    "month_ref",
    "cod_cliente",
    "nome_cliente",
    "cod_assessor",
    "nome_assessor",
    "to_email",
    "cc_email",
    "token",
    "subject",
    "entry_id",
    "conversation_id",
    "internet_message_id",
    "status",
    "last_update_at",
    "notes"
]


def _get_or_create_sheet(wb, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # se a aba existir mas estiver vazia, garante header
        if ws.max_row < 1:
            ws.append(DEFAULT_HEADERS)
            ws.freeze_panes = "A2"
        else:
            # se não tiver header, adiciona
            header = [c.value for c in ws[1]]
            if header != DEFAULT_HEADERS:
                # não sobrescreve automaticamente pra evitar bagunçar histórico antigo
                # apenas garante freeze panes se já tiver conteúdo
                ws.freeze_panes = "A2"
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(DEFAULT_HEADERS)
        ws.freeze_panes = "A2"
    return ws


def append_row(history_xlsx: str, sheet_name: str, row: dict):
    p = Path(history_xlsx)
    p.parent.mkdir(parents=True, exist_ok=True)

    if p.exists():
        wb = load_workbook(p)
    else:
        wb = Workbook()
        # remove "Sheet" padrão do openpyxl se existir
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    ws = _get_or_create_sheet(wb, sheet_name)

    values = [
        row.get("datetime_sent", ""),
        row.get("month_ref", ""),
        row.get("cod_cliente", ""),
        row.get("nome_cliente", ""),
        row.get("cod_assessor", ""),
        row.get("nome_assessor", ""),
        row.get("to_email", ""),
        row.get("cc_email", ""),
        row.get("token", ""),
        row.get("subject", ""),
        row.get("entry_id", ""),
        row.get("conversation_id", ""),
        row.get("internet_message_id", ""),
        row.get("status", ""),
        row.get("last_update_at", ""),
        row.get("notes", "")
    ]

    ws.append(values)
    wb.save(p)


def update_status_by_token(history_xlsx: str, sheet_name: str, token: str, new_status: str, notes: str = ""):
    p = Path(history_xlsx)
    if not p.exists():
        raise FileNotFoundError(f"Histórico não encontrado: {p.resolve()}")

    wb = load_workbook(p)
    ws = _get_or_create_sheet(wb, sheet_name)

    header = [c.value for c in ws[1]]

    if "token" not in header or "status" not in header:
        raise RuntimeError("Aba de histórico não contém colunas obrigatórias (token/status).")

    col_token = header.index("token") + 1
    col_status = header.index("status") + 1
    col_last = header.index("last_update_at") + 1 if "last_update_at" in header else None
    col_notes = header.index("notes") + 1 if "notes" in header else None

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    found = False
    for r in range(2, ws.max_row + 1):
        cur_token = str(ws.cell(r, col_token).value or "").strip()
        if cur_token == token:
            ws.cell(r, col_status).value = new_status
            if col_last:
                ws.cell(r, col_last).value = now
            if notes and col_notes:
                ws.cell(r, col_notes).value = notes
            found = True
            break

    wb.save(p)
    return found
