from openpyxl import load_workbook

from .history_store import update_status_by_token
from .outlook_client import OutlookClient


def _iter_inbox_items(ns):
    inbox = ns.GetDefaultFolder(6)  # 6 = Inbox
    items = inbox.Items
    try:
        items.Sort("[ReceivedTime]", True)  # mais recentes primeiro
    except Exception:
        pass
    return items


def _has_token_in_mail(item, token: str) -> bool:
    token = token.strip()
    if not token:
        return False

    try:
        subj = str(getattr(item, "Subject", "") or "")
        if token in subj:
            return True
    except Exception:
        pass

    try:
        body = str(getattr(item, "Body", "") or "")
        if token in body:
            return True
    except Exception:
        pass

    return False


def followup(cfg, logger):
    history_xlsx = cfg.get("paths.history_xlsx")
    sheet_name = cfg.get("history.sheet_name", "Performance_Audit_History")
    month_ref = cfg.get("project.month_ref", "")

    from_smtp = cfg.get("outlook.from_smtp")
    store_hint = cfg.get("outlook.store_hint", "riscos")
    oc = OutlookClient(from_smtp=from_smtp, store_hint=store_hint)

    wb = load_workbook(history_xlsx)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba de histórico não encontrada: {sheet_name}")

    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]

    col_token = header.index("token") + 1
    col_status = header.index("status") + 1
    col_month = header.index("month_ref") + 1

    items = _iter_inbox_items(oc.ns)

    checked = 0
    answered = 0
    rebilled = 0

    # limita varredura da caixa pra evitar ficar pesado
    scan_limit = 5000

    for r in range(2, ws.max_row + 1):
        token = str(ws.cell(r, col_token).value or "").strip()
        status = str(ws.cell(r, col_status).value or "").strip().upper()
        mref = str(ws.cell(r, col_month).value or "").strip()

        if month_ref and mref != month_ref:
            continue

        if status not in ["ENVIADO", "COBRADO"]:
            continue

        checked += 1
        found = False

        for i in range(1, min(items.Count, scan_limit) + 1):
            it = items.Item(i)
            if _has_token_in_mail(it, token):
                found = True
                break

        if found:
            answered += 1
            update_status_by_token(
                history_xlsx, sheet_name, token,
                "RESPONDIDO",
                notes="Resposta encontrada na Inbox via token."
            )
            logger.info(f"[RESPONDIDO] token={token}")
        else:
            rebilled += 1
            update_status_by_token(
                history_xlsx, sheet_name, token,
                "COBRADO",
                notes="Sem resposta detectada (token não encontrado na Inbox)."
            )
            logger.warn(f"[COBRADO] token={token} (sem resposta detectada)")

    logger.info("==== FOLLOW-UP RESUMO ====")
    logger.info(f"Registros verificados: {checked}")
    logger.info(f"Respondidos: {answered}")
    logger.info(f"Cobrados (sem resposta): {rebilled}")

